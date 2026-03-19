import os
from datetime import datetime
from io import BytesIO
from pathlib import Path
from uuid import uuid4

from flask import Flask, flash, redirect, render_template, request, send_file, url_for
from flask_login import LoginManager, UserMixin, current_user, login_required, login_user, logout_user
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import func, or_
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename
from openpyxl import Workbook, load_workbook

db = SQLAlchemy()
login_manager = LoginManager()
login_manager.login_view = "login"
ALLOWED_LOGO_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp", ".svg"}

def create_app():
    app = Flask(__name__, instance_relative_config=True)
    app.config["SECRET_KEY"] = os.getenv("SECRET_KEY", "change-me-now")
    app.config["SQLALCHEMY_DATABASE_URI"] = os.getenv("DATABASE_URL", "sqlite:///" + os.path.join(app.instance_path, "evolve.db"))
    app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
    app.config["MAX_CONTENT_LENGTH"] = 5 * 1024 * 1024
    os.makedirs(app.instance_path, exist_ok=True)
    os.makedirs(Path(app.root_path) / "static" / "uploads", exist_ok=True)
    db.init_app(app)
    login_manager.init_app(app)
    with app.app_context():
        db.create_all()
        if not User.query.filter_by(username="admin").first():
            u = User(username="admin", role="admin")
            u.set_password(os.getenv("DEFAULT_ADMIN_PASSWORD", "admin123!"))
            db.session.add(u)
            db.session.commit()
        if not AppSetting.query.first():
            db.session.add(AppSetting())
            db.session.commit()

    @app.context_processor
    def inject_branding():
        return {"app_settings": AppSetting.query.first()}

    register_routes(app)
    return app

class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    role = db.Column(db.String(30), default="admin", nullable=False)
    def set_password(self, raw): self.password_hash = generate_password_hash(raw)
    def check_password(self, raw): return check_password_hash(self.password_hash, raw)

class AppSetting(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    company_name = db.Column(db.String(150), default="Evolve Impianti Srls")
    logo_path = db.Column(db.String(255), default="")
    bolla_prefix = db.Column(db.String(20), default="BOL")

class Technician(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), unique=True, nullable=False)
    phone = db.Column(db.String(50), default="")
    notes = db.Column(db.String(255), default="")

class WarehouseItem(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(80), nullable=False)
    category = db.Column(db.String(50), nullable=False, default="materiale")
    description = db.Column(db.String(255), nullable=False)
    serialized = db.Column(db.Boolean, default=False, nullable=False)
    serial = db.Column(db.String(120), index=True, default="")
    quantity = db.Column(db.Integer, default=1, nullable=False)
    unit = db.Column(db.String(20), default="pz")
    min_stock = db.Column(db.Integer, default=0)
    notes = db.Column(db.String(255), default="")
    client_default = db.Column(db.String(120), default="")
    assigned_to = db.Column(db.Integer, db.ForeignKey("technician.id"))
    last_transfer_date = db.Column(db.String(40), default="")
    last_client = db.Column(db.String(120), default="")
    last_job = db.Column(db.String(120), default="")
    technician = db.relationship("Technician", backref="mobile_items")

class Tool(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(80), nullable=False)
    serial = db.Column(db.String(120), default="")
    description = db.Column(db.String(255), nullable=False)
    charge_value = db.Column(db.Float, default=0)
    status = db.Column(db.String(40), default="disponibile")
    notes = db.Column(db.String(255), default="")
    assigned_to = db.Column(db.Integer, db.ForeignKey("technician.id"))
    technician = db.relationship("Technician", backref="tools")

class Van(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    plate = db.Column(db.String(30), unique=True, nullable=False)
    model = db.Column(db.String(120), default="")
    status = db.Column(db.String(40), default="attivo")
    notes = db.Column(db.String(255), default="")
    assigned_to = db.Column(db.Integer, db.ForeignKey("technician.id"))
    technician = db.relationship("Technician", backref="vans")

class Charge(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    technician_id = db.Column(db.Integer, db.ForeignKey("technician.id"))
    description = db.Column(db.String(255), nullable=False)
    amount = db.Column(db.Float, default=0, nullable=False)
    status = db.Column(db.String(40), default="aperto")
    notes = db.Column(db.String(255), default="")
    technician = db.relationship("Technician", backref="charges")

class Transfer(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    bolla_no = db.Column(db.String(40), unique=True, nullable=False)
    transfer_type = db.Column(db.String(20), default="out", nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    technician_id = db.Column(db.Integer, db.ForeignKey("technician.id"))
    client = db.Column(db.String(120), default="")
    job = db.Column(db.String(120), default="")
    notes = db.Column(db.String(255), default="")
    technician = db.relationship("Technician", backref="transfers")

class TransferItem(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    transfer_id = db.Column(db.Integer, db.ForeignKey("transfer.id"), nullable=False)
    warehouse_item_id = db.Column(db.Integer, db.ForeignKey("warehouse_item.id"))
    category = db.Column(db.String(50), default="")
    code = db.Column(db.String(80), default="")
    description = db.Column(db.String(255), default="")
    serial = db.Column(db.String(120), default="")
    quantity = db.Column(db.Integer, default=1)
    unit = db.Column(db.String(20), default="pz")
    transfer = db.relationship("Transfer", backref="items")
    warehouse_item = db.relationship("WarehouseItem")

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

def settings_obj():
    return AppSetting.query.first()

def now_it():
    return datetime.now().strftime("%d/%m/%Y %H:%M")

def next_bolla_no():
    prefix = settings_obj().bolla_prefix if settings_obj() else "BOL"
    return f"{prefix}-{Transfer.query.count() + 1:05d}"

def allowed_logo(filename):
    return Path(filename).suffix.lower() in ALLOWED_LOGO_EXTENSIONS

def excel_response(wb, filename):
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return send_file(stream, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

def register_routes(app):
    @app.route("/")
    def home():
        return redirect(url_for("dashboard" if current_user.is_authenticated else "login"))

    @app.route("/login", methods=["GET", "POST"])
    def login():
        if request.method == "POST":
            user = User.query.filter(func.lower(User.username) == request.form.get("username","").strip().lower()).first()
            if user and user.check_password(request.form.get("password","")):
                login_user(user)
                return redirect(url_for("dashboard"))
            flash("Credenziali non valide.", "danger")
        return render_template("login.html", title="Accesso")

    @app.route("/logout")
    @login_required
    def logout():
        logout_user()
        return redirect(url_for("login"))

    @app.route("/settings", methods=["GET","POST"])
    @login_required
    def settings():
        s = settings_obj()
        if request.method == "POST":
            s.company_name = request.form.get("company_name","").strip() or "Evolve Impianti Srls"
            s.bolla_prefix = request.form.get("bolla_prefix","").strip() or "BOL"
            if request.form.get("remove_logo") == "1":
                s.logo_path = ""
            logo = request.files.get("logo_file")
            if logo and logo.filename:
                if not allowed_logo(logo.filename):
                    flash("Formato logo non supportato.", "danger")
                    return redirect(url_for("settings"))
                ext = Path(secure_filename(logo.filename)).suffix.lower()
                name = f"logo_{uuid4().hex}{ext}"
                save_path = Path(app.root_path) / "static" / "uploads" / name
                logo.save(save_path)
                s.logo_path = f"uploads/{name}"
            db.session.commit()
            flash("Impostazioni salvate.", "success")
            return redirect(url_for("settings"))
        return render_template("settings.html", settings_obj=s, title="Impostazioni")

    @app.route("/dashboard")
    @login_required
    def dashboard():
        stats = {
            "technicians": Technician.query.count(),
            "central_items": WarehouseItem.query.filter(WarehouseItem.assigned_to.is_(None)).count(),
            "mobile_items": WarehouseItem.query.filter(WarehouseItem.assigned_to.is_not(None)).count(),
            "open_charges": Charge.query.filter_by(status="aperto").count(),
            "transfers": Transfer.query.count(),
        }
        transfers = Transfer.query.order_by(Transfer.created_at.desc()).limit(10).all()
        return render_template("dashboard.html", stats=stats, transfers=transfers, title="Dashboard")

    @app.route("/technicians", methods=["GET","POST"])
    @login_required
    def technicians():
        if request.method == "POST":
            name = request.form.get("name","").strip()
            if name:
                db.session.add(Technician(name=name, phone=request.form.get("phone","").strip(), notes=request.form.get("notes","").strip()))
                db.session.commit()
                flash("Tecnico salvato.", "success")
            return redirect(url_for("technicians"))
        q = request.args.get("q","").strip()
        query = Technician.query
        if q:
            ilike = f"%{q}%"
            query = query.filter(or_(Technician.name.ilike(ilike), Technician.phone.ilike(ilike), Technician.notes.ilike(ilike)))
        technicians = query.order_by(Technician.name.asc()).limit(1000).all()
        return render_template("technicians.html", technicians=technicians, total_count=Technician.query.count(), q=q, title="Tecnici")

    @app.route("/warehouse", methods=["GET","POST"])
    @login_required
    def warehouse():
        if request.method == "POST":
            db.session.add(WarehouseItem(
                code=request.form.get("code","").strip(),
                category=request.form.get("category","materiale").strip(),
                description=request.form.get("description","").strip(),
                serialized=request.form.get("serialized")=="si",
                serial=request.form.get("serial","").strip(),
                quantity=int(request.form.get("quantity","1") or 1),
                unit=request.form.get("unit","pz").strip(),
                min_stock=int(request.form.get("min_stock","0") or 0),
                notes=request.form.get("notes","").strip(),
                client_default=request.form.get("client_default","").strip(),
            ))
            db.session.commit()
            flash("Materiale caricato a magazzino generale.", "success")
            return redirect(url_for("warehouse"))
        items = WarehouseItem.query.filter(WarehouseItem.assigned_to.is_(None)).order_by(WarehouseItem.id.desc()).all()
        return render_template("warehouse.html", items=items, title="Magazzino")

    @app.route("/transfers", methods=["GET","POST"])
    @login_required
    def transfers():
        technicians = Technician.query.order_by(Technician.name.asc()).limit(1000).all()
        central_items = WarehouseItem.query.filter(WarehouseItem.assigned_to.is_(None)).order_by(WarehouseItem.id.desc()).all()
        if request.method == "POST":
            tech_id = int(request.form.get("technician_id"))
            client = request.form.get("client","").strip()
            job = request.form.get("job","").strip()
            notes = request.form.get("notes","").strip()
            selected_ids = request.form.getlist("item_ids")
            raw_serials = request.form.get("serials","")
            serials = [x.strip() for x in raw_serials.replace(";", "\n").replace(",", "\n").splitlines() if x.strip()]
            item_map = {}
            for sid in selected_ids:
                item = WarehouseItem.query.get(int(sid))
                if item and item.assigned_to is None:
                    item_map[item.id] = item
            missing = []
            for serial in serials:
                item = WarehouseItem.query.filter(WarehouseItem.assigned_to.is_(None), func.lower(WarehouseItem.serial) == serial.lower()).first()
                if item:
                    item_map[item.id] = item
                else:
                    missing.append(serial)
            found = list(item_map.values())
            if not found:
                flash("Nessun materiale valido trovato.", "danger")
                return redirect(url_for("transfers"))
            if not client:
                client = next((i.client_default for i in found if i.client_default), "")
            if not client:
                flash("Inserisci la committente o imposta la committente predefinita sugli articoli.", "danger")
                return redirect(url_for("transfers"))
            tr = Transfer(bolla_no=next_bolla_no(), transfer_type="out", technician_id=tech_id, client=client, job=job, notes=notes)
            db.session.add(tr)
            db.session.flush()
            for item in found:
                item.assigned_to = tech_id
                item.last_transfer_date = now_it()
                item.last_client = client
                item.last_job = job
                db.session.add(TransferItem(transfer_id=tr.id, warehouse_item_id=item.id, category=item.category, code=item.code, description=item.description, serial=item.serial, quantity=item.quantity, unit=item.unit))
            db.session.commit()
            msg = f"Bolla {tr.bolla_no} creata con {len(found)} righe."
            if missing:
                msg += " Seriali non trovati: " + ", ".join(missing)
            flash(msg, "success")
            return redirect(url_for("transfer_detail", transfer_id=tr.id))
        transfers = Transfer.query.order_by(Transfer.created_at.desc()).limit(30).all()
        return render_template("transfers.html", technicians=technicians, central_items=central_items, transfers=transfers, title="Bolle")

    @app.route("/transfer/<int:transfer_id>")
    @login_required
    def transfer_detail(transfer_id):
        return render_template("transfer_detail.html", transfer=Transfer.query.get_or_404(transfer_id), title="Bolla")

    @app.route("/technician/<int:tech_id>")
    @login_required
    def technician_detail(tech_id):
        tech = Technician.query.get_or_404(tech_id)
        mobile_items = WarehouseItem.query.filter_by(assigned_to=tech.id).order_by(WarehouseItem.id.desc()).all()
        tools = Tool.query.filter_by(assigned_to=tech.id).all()
        van = Van.query.filter_by(assigned_to=tech.id).first()
        return render_template("technician_detail.html", tech=tech, mobile_items=mobile_items, tools=tools, van=van, title=f"Scheda {tech.name}")

    @app.route("/import/general", methods=["POST"])
    @login_required
    def import_general():
        file = request.files.get("file")
        if not file:
            flash("Seleziona un file Excel.", "danger")
            return redirect(url_for("warehouse"))
        wb = load_workbook(file)
        ws = wb["magazzino_generale"] if "magazzino_generale" in wb.sheetnames else wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            flash("File vuoto.", "danger")
            return redirect(url_for("warehouse"))
        headers = [str(x).strip() if x is not None else "" for x in rows[0]]
        imported = 0
        for row in rows[1:]:
            data = dict(zip(headers, row))
            code = str(data.get("codice","") or "").strip()
            desc = str(data.get("descrizione","") or "").strip()
            if not code or not desc:
                continue
            db.session.add(WarehouseItem(
                code=code,
                category=str(data.get("categoria","materiale") or "materiale").strip(),
                description=desc,
                serialized=str(data.get("serializzato","") or "").lower()=="si",
                serial=str(data.get("seriale","") or "").strip(),
                quantity=int(data.get("quantita",1) or 1),
                unit=str(data.get("unita","pz") or "pz").strip(),
                min_stock=int(data.get("scorta_minima",0) or 0),
                client_default=str(data.get("committente_predefinita","") or "").strip(),
                notes=str(data.get("note","") or "").strip(),
            ))
            imported += 1
        db.session.commit()
        flash(f"Importati {imported} materiali nel magazzino generale.", "success")
        return redirect(url_for("warehouse"))

    @app.route("/import/mobile", methods=["POST"])
    @login_required
    def import_mobile():
        file = request.files.get("file")
        if not file:
            flash("Seleziona un file Excel.", "danger")
            return redirect(url_for("transfers"))
        wb = load_workbook(file)
        ws = wb["magazzino_viaggiante"] if "magazzino_viaggiante" in wb.sheetnames else wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            flash("File vuoto.", "danger")
            return redirect(url_for("transfers"))
        headers = [str(x).strip() if x is not None else "" for x in rows[0]]
        imported = skipped = 0
        for row in rows[1:]:
            data = dict(zip(headers, row))
            tech_name = str(data.get("tecnico","") or "").strip()
            tech = Technician.query.filter(func.lower(Technician.name) == tech_name.lower()).first()
            code = str(data.get("codice","") or "").strip()
            desc = str(data.get("descrizione","") or "").strip()
            if not tech or not code or not desc:
                skipped += 1
                continue
            db.session.add(WarehouseItem(
                code=code,
                category=str(data.get("categoria","materiale") or "materiale").strip(),
                description=desc,
                serialized=str(data.get("serializzato","") or "").lower()=="si",
                serial=str(data.get("seriale","") or "").strip(),
                quantity=int(data.get("quantita",1) or 1),
                unit=str(data.get("unita","pz") or "pz").strip(),
                min_stock=int(data.get("scorta_minima",0) or 0),
                client_default=str(data.get("committente_predefinita","") or "").strip(),
                notes=str(data.get("note","") or "").strip(),
                assigned_to=tech.id,
                last_transfer_date=str(data.get("data_consegna","") or now_it()).strip(),
                last_client=str(data.get("committente","") or "").strip(),
                last_job=str(data.get("commessa","") or "").strip(),
            ))
            imported += 1
        db.session.commit()
        flash(f"Importati {imported} materiali ai tecnici. Scartate {skipped} righe.", "success")
        return redirect(url_for("transfers"))

    @app.route("/export/full")
    @login_required
    def export_full():
        wb = Workbook()
        ws = wb.active
        ws.title = "magazzino_generale"
        ws.append(["categoria","codice","descrizione","serializzato","seriale","quantita","unita","scorta_minima","committente_predefinita","note"])
        for item in WarehouseItem.query.filter(WarehouseItem.assigned_to.is_(None)).all():
            ws.append([item.category,item.code,item.description,"si" if item.serialized else "no",item.serial,item.quantity,item.unit,item.min_stock,item.client_default,item.notes])
        ws2 = wb.create_sheet("magazzino_viaggiante")
        ws2.append(["tecnico","categoria","codice","descrizione","serializzato","seriale","quantita","unita","scorta_minima","committente_predefinita","committente","commessa","data_consegna","note"])
        for item in WarehouseItem.query.filter(WarehouseItem.assigned_to.is_not(None)).all():
            ws2.append([item.technician.name if item.technician else "",item.category,item.code,item.description,"si" if item.serialized else "no",item.serial,item.quantity,item.unit,item.min_stock,item.client_default,item.last_client,item.last_job,item.last_transfer_date,item.notes])
        return excel_response(wb, "Evolve_Export_Completo.xlsx")

    @app.route("/export/general")
    @login_required
    def export_general():
        wb = Workbook()
        ws = wb.active
        ws.title = "magazzino_generale"
        ws.append(["categoria","codice","descrizione","serializzato","seriale","quantita","unita","scorta_minima","committente_predefinita","note"])
        for item in WarehouseItem.query.filter(WarehouseItem.assigned_to.is_(None)).all():
            ws.append([item.category,item.code,item.description,"si" if item.serialized else "no",item.serial,item.quantity,item.unit,item.min_stock,item.client_default,item.notes])
        return excel_response(wb, "Evolve_Magazzino_Generale.xlsx")

    @app.route("/export/mobile")
    @login_required
    def export_mobile():
        wb = Workbook()
        ws = wb.active
        ws.title = "magazzino_viaggiante"
        ws.append(["tecnico","categoria","codice","descrizione","serializzato","seriale","quantita","unita","scorta_minima","committente_predefinita","committente","commessa","data_consegna","note"])
        for item in WarehouseItem.query.filter(WarehouseItem.assigned_to.is_not(None)).all():
            ws.append([item.technician.name if item.technician else "",item.category,item.code,item.description,"si" if item.serialized else "no",item.serial,item.quantity,item.unit,item.min_stock,item.client_default,item.last_client,item.last_job,item.last_transfer_date,item.notes])
        return excel_response(wb, "Evolve_Magazzino_Viaggiante.xlsx")

app = create_app()

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
